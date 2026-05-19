import pandas as pd
import tkinter as tk
import io
import math

from tkinter import ttk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# =========================
# VARIABLES GLOBALES
# =========================

df = None
resultado = None

# =========================
# MOSTRAR TABLA
# =========================

def mostrar_tabla(dataframe):
    tree.delete(*tree.get_children())
    tree["columns"] = list(dataframe.columns)
    tree["show"] = "headings"
    for columna in dataframe.columns:
        tree.heading(
            columna,
            text=columna.upper(),
            command=lambda c=columna: ordenar_columna(c)
        )
        tree.column(columna, width=150, anchor="center")
    for _, fila in dataframe.iterrows():
        tree.insert("", "end", values=list(fila))


def ordenar_columna(columna):
    global df
    if df is None:
        return
    try:
        df = df.sort_values(by=columna)
        mostrar_tabla(df)
    except Exception:
        pass


# =========================
# CARGAR ARCHIVO
# =========================

def cargar_csv():
    global df
    ruta = filedialog.askopenfilename(
        title="Seleccionar archivo de datos",
        filetypes=[
            ("Todos los archivos de datos", "*.csv *.xlsx *.xls *.tsv *.txt"),
            ("CSV",   "*.csv"),
            ("Excel", "*.xlsx *.xls"),
            ("TSV",   "*.tsv"),
            ("Texto", "*.txt"),
            ("Todos", "*.*"),
        ]
    )
    if not ruta:
        return
    try:
        ext = ruta.rsplit(".", 1)[-1].lower()

        if ext in ("xlsx", "xls"):
            hojas = pd.ExcelFile(ruta).sheet_names
            if len(hojas) > 1:
                hoja = _elegir_hoja(hojas)
                if hoja is None:
                    return
            else:
                hoja = hojas[0]
            df = pd.read_excel(ruta, sheet_name=hoja)

        elif ext == "tsv":
            df = pd.read_csv(ruta, sep="\t", encoding="utf-8-sig")

        elif ext == "txt":
            with open(ruta, "r", encoding="utf-8-sig", errors="replace") as f:
                muestra = f.read(2048)
            sep = "\t" if muestra.count("\t") > muestra.count(",") else ","
            df = pd.read_csv(ruta, sep=sep, encoding="utf-8-sig")

        else:
            cargado = False
            for sep in (",", ";", "|", "\t"):
                try:
                    tmp = pd.read_csv(ruta, sep=sep, encoding="utf-8-sig")
                    if len(tmp.columns) > 1:
                        df = tmp
                        cargado = True
                        break
                except Exception:
                    continue
            if not cargado:
                df = pd.read_csv(ruta, encoding="utf-8-sig")

        limpiar_dataframe()
        mostrar_tabla(df)
        _actualizar_combos_busqueda_filtro()
        messagebox.showinfo(
            "Archivo cargado",
            f"Cargado correctamente\n\n"
            f"Filas: {len(df)}  |  Columnas: {len(df.columns)}\n"
            f"Columnas detectadas: {', '.join(df.columns)}"
        )
    except Exception as e:
        messagebox.showerror("Error al cargar", str(e))


def _elegir_hoja(hojas):
    win = tk.Toplevel(root)
    win.title("Seleccionar hoja")
    win.geometry("300x160")
    win.configure(bg="#1e1e1e")
    win.grab_set()
    elegida = tk.StringVar(value=hojas[0])
    tk.Label(
        win,
        text="El archivo tiene varias hojas.\nElige cual cargar:",
        bg="#1e1e1e", fg="white", font=("Segoe UI", 10)
    ).pack(pady=10)
    ttk.Combobox(
        win, textvariable=elegida, values=hojas,
        state="readonly", width=24
    ).pack(pady=5)
    resultado_hoja = [None]

    def confirmar():
        resultado_hoja[0] = elegida.get()
        win.destroy()

    ttk.Button(win, text="Cargar", command=confirmar).pack(pady=8)
    win.wait_window()
    return resultado_hoja[0]


# =========================
# PEGAR DATOS
# =========================

def cargar_desde_texto():
    global df
    contenido = input_text.get("1.0", tk.END).strip()
    try:
        if "," in contenido:
            df = pd.read_csv(io.StringIO(contenido), sep=",")
        else:
            df = pd.read_csv(io.StringIO(contenido), sep=r"\s+")
        limpiar_dataframe()
        mostrar_tabla(df)
        _actualizar_combos_busqueda_filtro()
        messagebox.showinfo("Exito", "Datos cargados correctamente")
    except Exception as e:
        messagebox.showerror("Error", "Formato invalido\n" + str(e))


# =========================
# LIMPIAR DATAFRAME
# =========================

def limpiar_dataframe():
    global df
    df.columns = df.columns.str.strip().str.lower()
    for columna in df.columns:
        if df[columna].dtype == "object":
            convertida = pd.to_numeric(df[columna], errors="coerce")
            if convertida.notna().sum() > len(df) * 0.6:
                df[columna] = convertida
            else:
                df[columna] = df[columna].astype(str).str.strip()


# =========================
# ACTUALIZAR COMBOS
# =========================

def _actualizar_combos_busqueda_filtro():
    if df is None:
        return
    cols_todas = list(df.columns)
    cols_num   = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

    combo_buscar_col["values"] = cols_todas
    if cols_todas:
        combo_buscar_col.set(cols_todas[0])

    combo_filtro_col["values"] = cols_num
    if cols_num:
        combo_filtro_col.set(cols_num[0])
    else:
        combo_filtro_col.set("")


# =========================
# VALIDAR
# =========================

def validar():
    if df is None:
        messagebox.showwarning("Aviso", "Primero carga datos")
        return
    vacios = df.isnull().sum()
    vacios = vacios[vacios > 0]
    if not vacios.empty:
        detalle = "\n".join(
            [f"  - {col}: {n} vacio(s)" for col, n in vacios.items()]
        )
        messagebox.showwarning("Validacion", f"Hay datos vacios:\n\n{detalle}")
    else:
        messagebox.showinfo(
            "Validacion",
            f"Datos validos\n\nRegistros: {len(df)}  |  Columnas: {len(df.columns)}"
        )


# =========================
# BUSCAR
# =========================

def buscar_tiempo_real(event=None):
    global resultado
    if df is None:
        return
    texto = entry_buscar.get().lower().strip()
    col   = combo_buscar_col.get()
    if not col or col not in df.columns:
        return
    if texto == "":
        mostrar_tabla(df)
        return
    resultado = df[
        df[col].astype(str).str.lower().str.contains(texto, na=False)
    ]
    mostrar_tabla(resultado)


# =========================
# FILTRAR
# =========================

def filtrar_tiempo_real(event=None):
    global resultado
    if df is None:
        return
    valor = entry_filtro_num.get().strip()
    col   = combo_filtro_col.get()
    if not col or col not in df.columns:
        return
    if valor == "":
        mostrar_tabla(df)
        return
    try:
        num = float(valor)
        resultado = df[pd.to_numeric(df[col], errors="coerce") >= num]
        mostrar_tabla(resultado)
    except Exception:
        pass


# =========================
# ESTADISTICAS
# =========================

def estadisticas():
    if df is None:
        messagebox.showwarning("Aviso", "Primero carga datos")
        return
    cols_num = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
    if not cols_num:
        messagebox.showwarning(
            "Estadisticas",
            "No hay columnas numericas en los datos cargados."
        )
        return
    try:
        lineas = [f"Estadisticas   {len(df)} registros\n"]
        for col in cols_num:
            serie = pd.to_numeric(df[col], errors="coerce").dropna()
            if serie.empty:
                continue
            lineas.append(
                f"-- {col.upper()} --\n"
                f"  Promedio : {serie.mean():.2f}\n"
                f"  Maximo   : {serie.max()}\n"
                f"  Minimo   : {serie.min()}\n"
                f"  Mediana  : {serie.median():.2f}\n"
                f"  Desv.std : {serie.std():.2f}\n"
            )
        messagebox.showinfo("Estadisticas", "\n".join(lineas))
    except Exception as e:
        messagebox.showerror("Error", str(e))


# =========================
# EXPORTAR
# =========================

def exportar():
    global resultado
    if resultado is None:
        messagebox.showwarning("Aviso", "No hay resultados para exportar")
        return
    ruta = filedialog.asksaveasfilename(defaultextension=".csv")
    if ruta:
        resultado.to_csv(ruta, index=False)
        messagebox.showinfo("Exito", "Archivo exportado correctamente")


# ==========================================================
# CONFIGURADOR DE GRAFICOS PRO — COMPLETAMENTE RENOVADO
# ==========================================================

def mostrar_grafico():
    if df is None:
        messagebox.showwarning("Aviso", "Primero carga datos")
        return
    VentanaGrafico(root, df)


# ── Paletas de colores profesionales ──────────────────────
PALETAS = {
    "Ocean Pro": [
        "#0077b6", "#00b4d8", "#90e0ef", "#023e8a", "#0096c7",
        "#ade8f4", "#48cae4", "#0077b6", "#00b4d8", "#90e0ef",
        "#023e8a", "#0096c7"
    ],
    "Sunset": [
        "#f72585", "#b5179e", "#7209b7", "#560bad", "#480ca8",
        "#3a0ca3", "#3f37c9", "#4361ee", "#4895ef", "#4cc9f0",
        "#f72585", "#b5179e"
    ],
    "Forest": [
        "#1b4332", "#2d6a4f", "#40916c", "#52b788", "#74c69d",
        "#95d5b2", "#b7e4c7", "#d8f3dc", "#1b4332", "#2d6a4f",
        "#40916c", "#52b788"
    ],
    "Corporate": [
        "#003049", "#d62828", "#f77f00", "#fcbf49", "#eae2b7",
        "#003049", "#d62828", "#f77f00", "#fcbf49", "#eae2b7",
        "#003049", "#d62828"
    ],
    "Neon Dark": [
        "#ff006e", "#8338ec", "#3a86ff", "#06d6a0", "#ffbe0b",
        "#fb5607", "#ff006e", "#8338ec", "#3a86ff", "#06d6a0",
        "#ffbe0b", "#fb5607"
    ],
}


class VentanaGrafico(tk.Toplevel):

    BG        = "#0d0d1a"
    BG2       = "#13132b"
    BG3       = "#1a1a3a"
    ACCENT    = "#2d2d6b"
    HIGHLIGHT = "#7c3aed"
    ACCENT2   = "#06d6a0"
    FG        = "#f1f1f5"
    FG2       = "#8888aa"
    SUCCESS   = "#06d6a0"
    WARNING   = "#ffbe0b"
    ERROR     = "#ff006e"

    TIPOS = {
        "📊 Barras"    : "bar",
        "📈 Lineal"    : "line",
        "🥧 Circular"  : "pie",
        "📉 Área"      : "area",
        "🔵 Dispersión": "scatter",
    }

    def __init__(self, parent, dataframe):
        super().__init__(parent)
        self.df = dataframe.copy()
        self.title("Chart Studio PRO  —  CSV Data Analyzer ")
        self.geometry("1160x740")
        self.minsize(1000, 680)
        self.configure(bg=self.BG)
        self.grab_set()

        # Variables de control
        self.tipo_var     = tk.StringVar(value="📊 Barras")
        self.col_x_var    = tk.StringVar()
        self.col_y_var    = tk.StringVar()
        self.titulo_var   = tk.StringVar()
        self.paleta_var   = tk.StringVar(value="Ocean Pro")
        self.mostrar_todo = tk.BooleanVar(value=True)
        self.max_items    = tk.IntVar(value=50)
        self.mostrar_vals = tk.BooleanVar(value=True)
        self.mostrar_grid = tk.BooleanVar(value=True)

        # Clasificación de columnas
        self.cols_all  = list(self.df.columns)
        self.cols_num  = [c for c in self.df.columns
                          if pd.api.types.is_numeric_dtype(self.df[c])]
        self.cols_text = [c for c in self.df.columns
                          if not pd.api.types.is_numeric_dtype(self.df[c])]

        self._build_ui()
        self._actualizar_opciones()
        self._previsualizar()

    # ── Construcción UI ───────────────────────────────────

    def _build_ui(self):
        # Cabecera
        header = tk.Frame(self, bg=self.BG, height=56)
        header.pack(fill="x", padx=0, pady=0)
        header.pack_propagate(False)

        tk.Label(
            header, text="  ◈  Chart Studio PRO",
            bg=self.BG, fg=self.FG,
            font=("Segoe UI", 15, "bold")
        ).pack(side="left", padx=20, pady=14)

        self.lbl_estado = tk.Label(
            header, text="", bg=self.BG, fg=self.SUCCESS,
            font=("Segoe UI", 9, "italic")
        )
        self.lbl_estado.pack(side="right", padx=20)

        tk.Frame(self, bg=self.ACCENT, height=1).pack(fill="x")

        # Cuerpo
        body = tk.Frame(self, bg=self.BG)
        body.pack(fill="both", expand=True)

        self._panel_lateral(body)
        self._panel_preview(body)

        # Barra inferior
        tk.Frame(self, bg=self.ACCENT, height=1).pack(fill="x")
        self._barra_inferior()

    def _panel_lateral(self, parent):
        """Panel izquierdo con todos los controles."""
        canal = tk.Frame(parent, bg=self.BG2, width=310)
        canal.pack(side="left", fill="y")
        canal.pack_propagate(False)

        canvas_lat = tk.Canvas(canal, bg=self.BG2, bd=0,
                               highlightthickness=0, width=310)
        scroll_lat = ttk.Scrollbar(canal, orient="vertical",
                                   command=canvas_lat.yview)
        canvas_lat.configure(yscrollcommand=scroll_lat.set)
        scroll_lat.pack(side="right", fill="y")
        canvas_lat.pack(side="left", fill="both", expand=True)

        self.panel = tk.Frame(canvas_lat, bg=self.BG2, width=290)
        canvas_lat.create_window((0, 0), window=self.panel, anchor="nw")
        self.panel.bind("<Configure>",
                        lambda e: canvas_lat.configure(
                            scrollregion=canvas_lat.bbox("all")))

        # --- Tipo de gráfico ---
        self._seccion("  TIPO DE GRÁFICO")
        tipo_grid = tk.Frame(self.panel, bg=self.BG2)
        tipo_grid.pack(padx=14, pady=(4, 8), anchor="w")
        for i, texto in enumerate(self.TIPOS):
            rb = tk.Radiobutton(
                tipo_grid, text=texto, variable=self.tipo_var,
                value=texto, bg=self.BG2, fg=self.FG,
                selectcolor=self.HIGHLIGHT,
                activebackground=self.BG2, activeforeground=self.ACCENT2,
                font=("Segoe UI", 10), cursor="hand2",
                command=self._on_tipo_change
            )
            rb.grid(row=i // 2, column=i % 2, sticky="w",
                    padx=6, pady=3)

        # --- Columna X ---
        self._seccion("  EJE X  (categoría / horizontal)")
        self._lbl("Cualquier columna disponible:")
        self.cb_x = self._combo(self.col_x_var, self.cols_all)
        self.cb_x.bind("<<ComboboxSelected>>",
                       lambda e: self._previsualizar())

        # --- Columna Y ---
        self._seccion("  EJE Y  (valor / vertical)")
        self._lbl("Cualquier columna disponible:")
        self.cb_y = self._combo(self.col_y_var, self.cols_all)
        self.cb_y.bind("<<ComboboxSelected>>",
                       lambda e: self._previsualizar())
        self._lbl_hint = self._lbl(
            "✦ Para barras/líneas se recomienda columna numérica en Y",
            color="#666688"
        )

        # --- Paleta de colores ---
        self._seccion("  PALETA DE COLORES")
        cb_pal = self._combo(self.paleta_var, list(PALETAS.keys()))
        cb_pal.bind("<<ComboboxSelected>>", lambda e: self._previsualizar())

        # --- Título ---
        self._seccion("  TÍTULO DEL GRÁFICO")
        e_titulo = tk.Entry(
            self.panel, textvariable=self.titulo_var,
            bg=self.BG3, fg=self.FG, insertbackground=self.FG,
            font=("Segoe UI", 10), relief="flat", bd=8, width=28
        )
        e_titulo.pack(padx=14, pady=(0, 10), anchor="w")
        e_titulo.bind("<KeyRelease>", lambda e: self._previsualizar())

        # --- Control de datos ---
        self._seccion("  CONTROL DE DATOS")

        fila_todo = tk.Frame(self.panel, bg=self.BG2)
        fila_todo.pack(padx=14, pady=(0, 6), anchor="w")
        tk.Checkbutton(
            fila_todo, text="Mostrar TODOS los registros",
            variable=self.mostrar_todo, bg=self.BG2, fg=self.FG,
            selectcolor=self.HIGHLIGHT, activebackground=self.BG2,
            font=("Segoe UI", 9), cursor="hand2",
            command=self._toggle_limite
        ).pack(side="left")

        self.frame_limite = tk.Frame(self.panel, bg=self.BG2)
        self.frame_limite.pack(padx=14, pady=(0, 4), anchor="w")
        self._lbl_en(self.frame_limite, "Límite de registros:", side="left")
        tk.Spinbox(
            self.frame_limite, from_=5, to=10000,
            textvariable=self.max_items, width=8,
            bg=self.BG3, fg=self.FG, font=("Segoe UI", 10),
            relief="flat", command=self._previsualizar
        ).pack(side="left", padx=6)
        self.frame_limite.pack_forget()  # oculto al inicio

        # --- Opciones visuales ---
        self._seccion("  OPCIONES VISUALES")
        for txt, var in [
            ("Mostrar valores en barras/puntos", self.mostrar_vals),
            ("Mostrar cuadrícula",               self.mostrar_grid),
        ]:
            tk.Checkbutton(
                self.panel, text=txt, variable=var,
                bg=self.BG2, fg=self.FG,
                selectcolor=self.HIGHLIGHT, activebackground=self.BG2,
                font=("Segoe UI", 9), cursor="hand2",
                command=self._previsualizar
            ).pack(padx=14, pady=2, anchor="w")

        # Indicador de registros
        tk.Frame(self.panel, bg=self.ACCENT, height=1).pack(
            fill="x", padx=14, pady=14
        )
        self.lbl_info = tk.Label(
            self.panel, text="", bg=self.BG2, fg=self.FG2,
            font=("Segoe UI", 8, "italic"), wraplength=270, justify="left"
        )
        self.lbl_info.pack(padx=14, pady=(0, 14), anchor="w")

    def _panel_preview(self, parent):
        """Panel derecho con la vista previa del gráfico."""
        import matplotlib
        matplotlib.use("TkAgg")
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
        import matplotlib.pyplot as plt

        self._Figure = Figure
        self._Canvas = FigureCanvasTkAgg
        self._plt    = plt

        panel = tk.Frame(parent, bg=self.BG)
        panel.pack(side="left", fill="both", expand=True)

        # Mini toolbar
        toolbar = tk.Frame(panel, bg=self.BG3, height=34)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)
        tk.Label(
            toolbar, text="VISTA PREVIA",
            bg=self.BG3, fg=self.FG2, font=("Segoe UI", 8, "bold")
        ).pack(side="left", padx=16, pady=8)

        btn_refresh = tk.Button(
            toolbar, text="↺  Actualizar",
            bg=self.BG3, fg=self.ACCENT2,
            font=("Segoe UI", 8, "bold"), relief="flat",
            cursor="hand2", bd=0, pady=0,
            command=self._previsualizar
        )
        btn_refresh.pack(side="right", padx=12, pady=6)

        # Canvas matplotlib
        area = tk.Frame(panel, bg=self.BG)
        area.pack(fill="both", expand=True, padx=12, pady=8)

        self.fig = self._Figure(figsize=(7.8, 5.8), dpi=96,
                                facecolor=self.BG)
        self.canvas_mpl = self._Canvas(self.fig, master=area)
        self.canvas_mpl.get_tk_widget().pack(fill="both", expand=True)

    def _barra_inferior(self):
        barra = tk.Frame(self, bg=self.BG2, height=58)
        barra.pack(fill="x")
        barra.pack_propagate(False)

        # Botón principal
        tk.Button(
            barra, text="  ⬇  Exportar a Excel con gráfico",
            bg=self.HIGHLIGHT, fg="white",
            font=("Segoe UI", 11, "bold"), relief="flat",
            padx=20, pady=10, cursor="hand2",
            command=self._generar_excel
        ).pack(side="left", padx=16, pady=8)

        tk.Button(
            barra, text="  ↺  Actualizar preview",
            bg=self.ACCENT, fg=self.FG,
            font=("Segoe UI", 10), relief="flat",
            padx=14, pady=10, cursor="hand2",
            command=self._previsualizar
        ).pack(side="left", padx=4, pady=8)

        tk.Button(
            barra, text="  ✕  Cerrar",
            bg=self.BG3, fg=self.FG2,
            font=("Segoe UI", 10), relief="flat",
            padx=14, pady=10, cursor="hand2",
            command=self.destroy
        ).pack(side="right", padx=16, pady=8)

    # ── Helpers UI ────────────────────────────────────────

    def _seccion(self, texto):
        tk.Label(
            self.panel, text=texto,
            bg=self.BG2, fg=self.ACCENT2,
            font=("Segoe UI", 8, "bold")
        ).pack(anchor="w", padx=14, pady=(14, 2))
        tk.Frame(self.panel, bg=self.ACCENT, height=1).pack(
            fill="x", padx=14, pady=(0, 6)
        )

    def _lbl(self, texto, color=None):
        lbl = tk.Label(
            self.panel, text=texto,
            bg=self.BG2, fg=color or self.FG2,
            font=("Segoe UI", 8), wraplength=270, justify="left"
        )
        lbl.pack(anchor="w", padx=14, pady=(0, 2))
        return lbl

    def _lbl_en(self, parent, texto, side="top"):
        tk.Label(
            parent, text=texto, bg=self.BG2, fg=self.FG2,
            font=("Segoe UI", 9)
        ).pack(side=side, padx=(0, 4))

    def _combo(self, var, valores, ancho=28):
        style_name = "Dark.TCombobox"
        cb = ttk.Combobox(
            self.panel, textvariable=var, values=valores,
            state="readonly", width=ancho, font=("Segoe UI", 10)
        )
        cb.pack(padx=14, pady=(0, 6), anchor="w")
        return cb

    # ── Lógica de controles ───────────────────────────────

    def _toggle_limite(self):
        if self.mostrar_todo.get():
            self.frame_limite.pack_forget()
        else:
            self.frame_limite.pack(padx=14, pady=(0, 4), anchor="w")
        self._previsualizar()

    def _on_tipo_change(self):
        self._actualizar_opciones()
        self._previsualizar()

    def _actualizar_opciones(self):
        """
        Ambos ejes permiten CUALQUIER columna.
        Solo para gráfico circular, Y se deshabilita.
        """
        tipo = self.TIPOS[self.tipo_var.get()]

        if tipo == "pie":
            self.cb_x["values"] = self.cols_all
            self.cb_y["state"] = "disabled"
            self.cb_y.set("")
        elif tipo == "scatter":
            # Scatter: ambos ejes idealmente numéricos, pero libre
            self.cb_x["values"] = self.cols_all
            self.cb_y["state"] = "readonly"
            self.cb_y["values"] = self.cols_all
        else:
            self.cb_x["values"] = self.cols_all
            self.cb_y["state"] = "readonly"
            self.cb_y["values"] = self.cols_all   # ← TODAS las columnas en Y

        # Valores por defecto inteligentes
        if not self.col_x_var.get() and self.cols_all:
            self.col_x_var.set(self.cols_all[0])
        if tipo != "pie" and not self.col_y_var.get():
            # Preferir columna numérica si existe
            if self.cols_num:
                self.col_y_var.set(self.cols_num[0])
            elif len(self.cols_all) > 1:
                self.col_y_var.set(self.cols_all[1])

    # ── Preparación de datos ──────────────────────────────

    def _get_datos(self):
        """
        Devuelve el subconjunto de datos según configuración.
        Sin límite si 'mostrar_todo' está activo.
        """
        if self.mostrar_todo.get():
            return self.df
        else:
            n = self.max_items.get()
            return self.df.head(n)

    def _validar(self):
        tipo  = self.TIPOS[self.tipo_var.get()]
        col_x = self.col_x_var.get()
        col_y = self.col_y_var.get()

        if not col_x or col_x not in self.df.columns:
            return False, "Selecciona una columna para el Eje X.", None

        if tipo in ("bar", "line", "area", "scatter"):
            if not col_y or col_y not in self.df.columns:
                return False, "Selecciona una columna para el Eje Y.", None
            if col_x == col_y:
                return False, "Eje X y Eje Y deben ser columnas distintas.", None

        if tipo == "pie":
            if len(self.df[col_x].dropna().unique()) < 2:
                return False, "La columna necesita al menos 2 categorías únicas.", None

        return True, "", {"tipo": tipo, "col_x": col_x, "col_y": col_y}

    # ── Vista previa matplotlib ───────────────────────────

    def _previsualizar(self, *_):
        ok, msg, datos = self._validar()

        # Estado
        if ok:
            self.lbl_estado.config(text="✔  Configuración válida", fg=self.SUCCESS)
        else:
            self.lbl_estado.config(text=f"⚠  {msg}", fg=self.WARNING)

        self.fig.clear()

        # Estilo global matplotlib
        with self._plt.style.context("dark_background"):
            ax = self.fig.add_subplot(111)

        ax.set_facecolor(self.BG3)
        self.fig.patch.set_facecolor(self.BG)

        if not ok:
            ax.text(0.5, 0.5, msg, transform=ax.transAxes,
                    ha="center", va="center",
                    color=self.WARNING, fontsize=12,
                    fontfamily="Segoe UI")
            ax.axis("off")
            self.canvas_mpl.draw()
            return

        tipo   = datos["tipo"]
        col_x  = datos["col_x"]
        col_y  = datos["col_y"]
        titulo = (self.titulo_var.get().strip()
                  or self._titulo_auto(tipo, col_x, col_y))
        paleta = PALETAS[self.paleta_var.get()]
        data   = self._get_datos()

        try:
            if tipo == "bar":
                self._grafico_barras(ax, data, col_x, col_y, paleta)
            elif tipo == "line":
                self._grafico_lineal(ax, data, col_x, col_y, paleta)
            elif tipo == "area":
                self._grafico_area(ax, data, col_x, col_y, paleta)
            elif tipo == "pie":
                self._grafico_circular(ax, data, col_x, paleta)
            elif tipo == "scatter":
                self._grafico_dispersion(ax, data, col_x, col_y, paleta)

            # Estilo compartido de ejes (excepto pie)
            if tipo != "pie":
                ax.set_title(titulo, color=self.FG, fontsize=13,
                             fontweight="bold", pad=16,
                             fontfamily="Segoe UI")
                ax.tick_params(colors=self.FG2, labelsize=8)
                for spine in ax.spines.values():
                    spine.set_color(self.ACCENT)
                if self.mostrar_grid.get():
                    ax.grid(axis="y", color=self.ACCENT,
                            alpha=0.4, linestyle="--", linewidth=0.7)
                ax.set_axisbelow(True)
            else:
                ax.set_title(titulo, color=self.FG, fontsize=13,
                             fontweight="bold", pad=16,
                             fontfamily="Segoe UI")

            # Info de registros
            n_real = len(data)
            n_total = len(self.df)
            sufijo = (f"  (mostrando {n_real} de {n_total})"
                      if n_real < n_total else "")
            self.lbl_info.config(
                text=f"Registros graficados: {n_real}{sufijo}\n"
                     f"Col X: {col_x}  •  "
                     + (f"Col Y: {col_y}" if col_y else ""),
                fg=self.SUCCESS
            )

            self.fig.tight_layout(pad=1.8)

        except Exception as e:
            ax.clear()
            ax.text(0.5, 0.5, f"Error al graficar:\n{e}",
                    transform=ax.transAxes, ha="center", va="center",
                    color=self.ERROR, fontsize=10)
            ax.axis("off")
            self.lbl_info.config(text=f"Error: {e}", fg=self.ERROR)

        self.canvas_mpl.draw()

    # ── Gráficos individuales ─────────────────────────────

    def _grafico_barras(self, ax, data, col_x, col_y, paleta):
        """
        Barras completas con agrupación inteligente para grandes datasets.
        Si hay muchos valores únicos en X, agrupa por frecuencia/suma.
        """
        sub = data[[col_x, col_y]].dropna().copy()

        # Convertir Y a numérico si es posible
        sub[col_y] = pd.to_numeric(sub[col_y], errors="coerce")
        sub = sub.dropna(subset=[col_y])

        n_unicos = sub[col_x].nunique()

        if n_unicos > 80:
            # Agrupación: suma por categoría, top 60
            sub = (sub.groupby(col_x, as_index=False)[col_y]
                   .sum()
                   .sort_values(col_y, ascending=False)
                   .head(60))
            nota = f" (top 60 de {n_unicos} categorías)"
        elif n_unicos > 30:
            sub = (sub.groupby(col_x, as_index=False)[col_y]
                   .sum()
                   .sort_values(col_y, ascending=False))
            nota = f" ({n_unicos} categorías)"
        else:
            sub = sub.groupby(col_x, as_index=False)[col_y].sum()
            nota = ""

        n = len(sub)
        colores_bar = (paleta * math.ceil(n / len(paleta)))[:n]

        bars = ax.bar(
            range(n), sub[col_y],
            color=colores_bar, edgecolor=self.BG, linewidth=0.6,
            zorder=3
        )

        ax.set_xticks(range(n))
        etiquetas = sub[col_x].astype(str).tolist()

        # Rotar etiquetas según cantidad
        if n > 20:
            ax.set_xticklabels(etiquetas, rotation=90,
                               ha="center", fontsize=6)
        elif n > 10:
            ax.set_xticklabels(etiquetas, rotation=45,
                               ha="right", fontsize=7)
        else:
            ax.set_xticklabels(etiquetas, rotation=20,
                               ha="right", fontsize=8)

        ax.set_xlabel(col_x.upper(), color=self.FG2, fontsize=9, labelpad=8)
        ax.set_ylabel(col_y.upper(), color=self.FG2, fontsize=9)

        # Valores encima de barras (solo si no son demasiadas)
        if self.mostrar_vals.get() and n <= 40:
            for bar, val in zip(bars, sub[col_y]):
                h = bar.get_height()
                ax.text(
                    bar.get_x() + bar.get_width() / 2,
                    h + (sub[col_y].max() * 0.01),
                    f"{val:,.1f}" if val != int(val) else f"{int(val):,}",
                    ha="center", va="bottom",
                    color=self.FG, fontsize=6, zorder=4
                )

        ax.set_title(
            ax.get_title() + nota if ax.get_title() else nota,
            color=self.FG2, fontsize=8
        )

    def _grafico_lineal(self, ax, data, col_x, col_y, paleta):
        """
        Gráfico lineal con TODOS los puntos, escalado inteligente.
        """
        sub = data[[col_x, col_y]].dropna().copy()
        sub[col_y] = pd.to_numeric(sub[col_y], errors="coerce")
        sub = sub.dropna(subset=[col_y]).reset_index(drop=True)

        n = len(sub)
        x_vals = range(n)

        ax.plot(
            x_vals, sub[col_y],
            color=paleta[0], linewidth=2,
            marker="o" if n <= 100 else None,
            markersize=4 if n <= 200 else 2,
            markerfacecolor=paleta[1],
            markeredgecolor=self.BG, markeredgewidth=0.5,
            zorder=3
        )
        ax.fill_between(x_vals, sub[col_y],
                        alpha=0.12, color=paleta[0], zorder=2)

        # Etiquetas en eje X (muestreadas si son muchas)
        etiquetas = sub[col_x].astype(str).tolist()
        paso = max(1, n // 20)
        ticks_pos = list(range(0, n, paso))
        ax.set_xticks(ticks_pos)
        ax.set_xticklabels(
            [etiquetas[i] for i in ticks_pos],
            rotation=45, ha="right",
            fontsize=7 if n > 30 else 8
        )

        ax.set_xlabel(col_x.upper(), color=self.FG2, fontsize=9, labelpad=8)
        ax.set_ylabel(col_y.upper(), color=self.FG2, fontsize=9)
        ax.set_xlim(-0.5, n - 0.5)

        # Puntos de min/max destacados
        if n > 1:
            idx_max = sub[col_y].idxmax()
            idx_min = sub[col_y].idxmin()
            ax.scatter([idx_max], [sub[col_y][idx_max]],
                       color=self.SUCCESS, s=60, zorder=5,
                       label=f"Máx: {sub[col_y][idx_max]:,.1f}")
            ax.scatter([idx_min], [sub[col_y][idx_min]],
                       color=self.ERROR, s=60, zorder=5,
                       label=f"Mín: {sub[col_y][idx_min]:,.1f}")
            ax.legend(facecolor=self.BG3, edgecolor=self.ACCENT,
                      labelcolor=self.FG, fontsize=8)

        self.lbl_info.config(
            text=f"Puntos graficados: {n}  •  "
                 f"Rango Y: {sub[col_y].min():,.1f} — {sub[col_y].max():,.1f}"
        )

    def _grafico_area(self, ax, data, col_x, col_y, paleta):
        """Gráfico de área rellena."""
        sub = data[[col_x, col_y]].dropna().copy()
        sub[col_y] = pd.to_numeric(sub[col_y], errors="coerce")
        sub = sub.dropna(subset=[col_y]).reset_index(drop=True)

        n = len(sub)
        x_vals = range(n)

        ax.fill_between(x_vals, sub[col_y],
                        color=paleta[0], alpha=0.55, zorder=2)
        ax.plot(x_vals, sub[col_y],
                color=paleta[0], linewidth=1.8, zorder=3)

        etiquetas = sub[col_x].astype(str).tolist()
        paso = max(1, n // 18)
        ticks_pos = list(range(0, n, paso))
        ax.set_xticks(ticks_pos)
        ax.set_xticklabels(
            [etiquetas[i] for i in ticks_pos],
            rotation=45, ha="right", fontsize=7
        )
        ax.set_xlabel(col_x.upper(), color=self.FG2, fontsize=9, labelpad=8)
        ax.set_ylabel(col_y.upper(), color=self.FG2, fontsize=9)
        ax.set_xlim(-0.5, n - 0.5)

    def _grafico_circular(self, ax, data, col_x, paleta):
        """
        Circular con porcentajes reales, leyenda lateral para muchas categorías.
        """
        conteo = data[col_x].dropna().value_counts()
        n_cats = len(conteo)

        # Si hay muchas categorías, agrupar las menores en "Otros"
        MAX_SLICES = 10
        if n_cats > MAX_SLICES:
            top    = conteo.head(MAX_SLICES - 1)
            otros  = conteo.iloc[MAX_SLICES - 1:].sum()
            conteo = pd.concat([top, pd.Series({"Otros": otros})])

        n = len(conteo)
        colores = (paleta * math.ceil(n / len(paleta)))[:n]

        wedges, texts, autotexts = ax.pie(
            conteo.values,
            labels=None,            # etiquetas en leyenda
            autopct="%1.1f%%",
            colors=colores,
            startangle=140,
            pctdistance=0.80,
            wedgeprops={
                "edgecolor": self.BG, "linewidth": 2,
                "antialiased": True
            },
            shadow=False
        )
        for at in autotexts:
            at.set_color(self.FG)
            at.set_fontsize(8)
            at.set_fontweight("bold")

        # Leyenda lateral
        ax.legend(
            wedges,
            [f"{lbl}  ({cnt:,})" for lbl, cnt
             in zip(conteo.index, conteo.values)],
            loc="center left",
            bbox_to_anchor=(1.02, 0.5),
            facecolor=self.BG3, edgecolor=self.ACCENT,
            labelcolor=self.FG, fontsize=8
        )

        self.lbl_info.config(
            text=f"Categorías: {n_cats}  •  "
                 f"Total registros: {conteo.sum():,}"
        )

    def _grafico_dispersion(self, ax, data, col_x, col_y, paleta):
        """Scatter plot con densidad de color."""
        sub = data[[col_x, col_y]].dropna().copy()
        sub[col_x] = pd.to_numeric(sub[col_x], errors="coerce")
        sub[col_y] = pd.to_numeric(sub[col_y], errors="coerce")
        sub = sub.dropna()

        n = len(sub)
        tam = max(8, min(60, 3000 // max(n, 1)))

        ax.scatter(
            sub[col_x], sub[col_y],
            color=paleta[0], alpha=0.6, s=tam,
            edgecolors=self.BG, linewidths=0.4, zorder=3
        )

        # Línea de tendencia (regresión simple)
        if n > 2:
            try:
                m, b = pd.Series(sub[col_y].values).corr(
                    pd.Series(sub[col_x].values)), 0
                import numpy as np
                coef = np.polyfit(sub[col_x], sub[col_y], 1)
                x_line = [sub[col_x].min(), sub[col_x].max()]
                y_line = [coef[0] * x + coef[1] for x in x_line]
                ax.plot(x_line, y_line, color=paleta[2],
                        linewidth=1.6, linestyle="--",
                        label="Tendencia", zorder=4)
                corr = sub[col_x].corr(sub[col_y])
                ax.legend(facecolor=self.BG3, edgecolor=self.ACCENT,
                          labelcolor=self.FG, fontsize=8,
                          title=f"r = {corr:.3f}",
                          title_fontsize=8)
            except Exception:
                pass

        ax.set_xlabel(col_x.upper(), color=self.FG2, fontsize=9, labelpad=8)
        ax.set_ylabel(col_y.upper(), color=self.FG2, fontsize=9)

    # ── Título automático ─────────────────────────────────

    def _titulo_auto(self, tipo, col_x, col_y):
        nombres = {
            "bar"    : f"{col_y.title()} por {col_x.title()}",
            "line"   : f"Tendencia de {col_y.title()} — {col_x.title()}",
            "area"   : f"Área: {col_y.title()} sobre {col_x.title()}",
            "pie"    : f"Distribución de {col_x.title()}",
            "scatter": f"Dispersión: {col_x.title()} vs {col_y.title()}",
        }
        return nombres.get(tipo, "Gráfico de análisis")

    # ── Exportación a Excel profesional ──────────────────

    def _generar_excel(self):
        """
        Exporta a Excel incrustando el gráfico matplotlib EXACTAMENTE
        como se ve en pantalla (imagen PNG de alta resolución),
        más hoja de datos y hoja de estadísticas con formato profesional.
        """
        ok, msg, datos = self._validar()
        if not ok:
            messagebox.showwarning("Configuración incompleta", msg, parent=self)
            return

        ruta = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Guardar análisis en Excel"
        )
        if not ruta:
            return

        try:
            import tempfile, os
            from openpyxl.drawing.image import Image as XLImage
            from matplotlib.figure import Figure
            from matplotlib.backends.backend_agg import FigureCanvasAgg

            tipo   = datos["tipo"]
            col_x  = datos["col_x"]
            col_y  = datos["col_y"]
            titulo = (self.titulo_var.get().strip()
                      or self._titulo_auto(tipo, col_x, col_y))
            data   = self._get_datos()
            paleta = PALETAS[self.paleta_var.get()]

            # ── 1. Renderizar el gráfico igual que en pantalla ─────────────
            #    Se crea una figura nueva de alta resolución (no la del canvas)
            fig_exp = Figure(figsize=(14, 9), dpi=150, facecolor=self.BG)
            canvas_exp = FigureCanvasAgg(fig_exp)
            ax_exp = fig_exp.add_subplot(111)
            ax_exp.set_facecolor(self.BG3)

            # Reutilizar exactamente las mismas funciones de dibujo
            if tipo == "bar":
                self._grafico_barras(ax_exp, data, col_x, col_y, paleta)
            elif tipo == "line":
                self._grafico_lineal(ax_exp, data, col_x, col_y, paleta)
            elif tipo == "area":
                self._grafico_area(ax_exp, data, col_x, col_y, paleta)
            elif tipo == "pie":
                self._grafico_circular(ax_exp, data, col_x, paleta)
            elif tipo == "scatter":
                self._grafico_dispersion(ax_exp, data, col_x, col_y, paleta)

            # Aplicar estilo idéntico al de la vista previa
            if tipo != "pie":
                ax_exp.set_title(titulo, color=self.FG, fontsize=18,
                                 fontweight="bold", pad=20)
                ax_exp.tick_params(colors=self.FG2, labelsize=10)
                for spine in ax_exp.spines.values():
                    spine.set_color(self.ACCENT)
                if self.mostrar_grid.get():
                    ax_exp.grid(axis="y", color=self.ACCENT,
                                alpha=0.4, linestyle="--", linewidth=0.8)
                ax_exp.set_axisbelow(True)
            else:
                ax_exp.set_title(titulo, color=self.FG, fontsize=18,
                                 fontweight="bold", pad=20)

            fig_exp.tight_layout(pad=2.0)

            # Guardar imagen en archivo temporal PNG
            tmp_png = tempfile.NamedTemporaryFile(
                suffix=".png", delete=False
            )
            tmp_png.close()
            fig_exp.savefig(
                tmp_png.name,
                dpi=150,
                facecolor=self.BG,
                bbox_inches="tight"
            )

            # ── 2. Construir el Excel ──────────────────────────────────────
            wb = Workbook()

            # Hoja 1: Gráfico (primera hoja, la más importante)
            ws_chart = wb.active
            ws_chart.title = "Gráfico"
            ws_chart.sheet_view.showGridLines = False
            ws_chart.sheet_properties.tabColor = "7c3aed"

            # Fondo oscuro para la hoja del gráfico
            fill_bg = PatternFill("solid", fgColor="0d0d1a")
            for row in ws_chart.iter_rows(min_row=1, max_row=60,
                                          min_col=1, max_col=30):
                for cell in row:
                    cell.fill = fill_bg

            # Título en la celda A1
            ws_chart["A1"] = titulo
            ws_chart["A1"].font = Font(
                bold=True, size=16,
                color="06d6a0", name="Segoe UI"
            )
            ws_chart["A1"].alignment = Alignment(
                horizontal="center", vertical="center"
            )
            ws_chart.row_dimensions[1].height = 32
            ws_chart.merge_cells("A1:P1")

            # Subtítulo con info
            n_registros = len(data)
            ws_chart["A2"] = (
                f"Tipo: {tipo.upper()}  •  "
                f"Columna X: {col_x.upper()}"
                + (f"  •  Columna Y: {col_y.upper()}" if col_y else "")
                + f"  •  Registros: {n_registros:,}"
            )
            ws_chart["A2"].font = Font(
                size=10, color="8888aa",
                italic=True, name="Segoe UI"
            )
            ws_chart["A2"].alignment = Alignment(horizontal="center")
            ws_chart.merge_cells("A2:P2")
            ws_chart.row_dimensions[2].height = 20

            # Insertar imagen PNG (gráfico matplotlib exacto)
            img_xl = XLImage(tmp_png.name)
            # Ajustar tamaño en puntos (ancho x alto en EMU internamente)
            img_xl.width  = 900   # píxeles aprox en Excel
            img_xl.height = 560
            ws_chart.add_image(img_xl, "A3")

            # Ajustar columnas para que la imagen quepa bien
            for col_letra in "ABCDEFGHIJKLMNOP":
                ws_chart.column_dimensions[col_letra].width = 9

            # Hoja 2: Dataset completo
            ws_data = wb.create_sheet("Dataset Completo")
            ws_data.sheet_properties.tabColor = "0077b6"
            self._estilizar_hoja_datos(ws_data, self.df)

            # Hoja 3: Datos usados en el gráfico
            ws_gdata = wb.create_sheet("Datos del Gráfico")
            ws_gdata.sheet_properties.tabColor = "00b4d8"
            self._escribir_datos_usados(ws_gdata, data, tipo,
                                        col_x, col_y, titulo)

            # Hoja 4: Resumen estadístico
            ws_stats = wb.create_sheet("Resumen Estadístico")
            ws_stats.sheet_properties.tabColor = "06d6a0"
            self._escribir_estadisticas(ws_stats)

            wb.save(ruta)

            # Limpiar PNG temporal
            try:
                os.unlink(tmp_png.name)
            except Exception:
                pass

            messagebox.showinfo(
                "Excel generado con éxito ✔",
                f"Archivo guardado correctamente.\n\n"
                f"  📊  Tipo de gráfico : {tipo.upper()}\n"
                f"  📋  Registros       : {n_registros:,}\n"
                f"  🎨  Paleta          : {self.paleta_var.get()}\n"
                f"  📁  Ruta            : {ruta}\n\n"
                f"Hojas incluidas:\n"
                f"  ➊  Gráfico  (imagen idéntica a la vista previa)\n"
                f"  ➋  Dataset Completo\n"
                f"  ➌  Datos del Gráfico\n"
                f"  ➍  Resumen Estadístico",
                parent=self
            )

        except Exception as e:
            messagebox.showerror("Error al generar Excel", str(e), parent=self)

    def _escribir_datos_usados(self, ws, data, tipo, col_x, col_y, titulo):
        """Escribe los datos exactos usados en el gráfico, con formato."""
        from openpyxl.utils import get_column_letter

        fill_h  = PatternFill("solid", fgColor="2d2d6b")
        font_h  = Font(bold=True, color="06d6a0", size=11, name="Segoe UI")
        alin_c  = Alignment(horizontal="center", vertical="center")
        fill_p  = PatternFill("solid", fgColor="13132b")
        fill_i  = PatternFill("solid", fgColor="1a1a3a")
        font_d  = Font(color="eaeaea", size=10, name="Segoe UI")
        borde   = Border(
            bottom=Side(style="thin", color="2d2d6b"),
            right=Side(style="thin", color="2d2d6b")
        )

        ws.sheet_view.showGridLines = False

        if tipo == "pie":
            conteo = data[col_x].dropna().value_counts()
            if len(conteo) > 10:
                top   = conteo.head(9)
                otros = conteo.iloc[9:].sum()
                conteo = pd.concat([top, pd.Series({"Otros": otros})])
            encabezados = [col_x.upper(), "CANTIDAD", "PORCENTAJE"]
            total = conteo.sum()
            filas = [
                [str(lbl), int(cnt), f"{cnt/total*100:.1f}%"]
                for lbl, cnt in conteo.items()
            ]
        else:
            sub = data[[col_x, col_y]].dropna().copy()
            sub[col_y] = pd.to_numeric(sub[col_y], errors="coerce")
            sub = sub.dropna(subset=[col_y])
            n_unicos = sub[col_x].nunique()
            if n_unicos > 80:
                sub = (sub.groupby(col_x, as_index=False)[col_y]
                       .sum().sort_values(col_y, ascending=False).head(60))
            elif n_unicos > 1:
                sub = sub.groupby(col_x, as_index=False)[col_y].sum()
            encabezados = [col_x.upper(), col_y.upper()]
            filas = list(sub.itertuples(index=False))

        # Título de sección
        ws["A1"] = f"Datos usados — {titulo}"
        ws["A1"].font = Font(bold=True, size=13, color="7c3aed",
                             name="Segoe UI")
        ws["A1"].alignment = alin_c
        ws.merge_cells(f"A1:{get_column_letter(len(encabezados))}1")
        ws.row_dimensions[1].height = 28

        # Encabezados
        for ci, enc in enumerate(encabezados, 1):
            c = ws.cell(row=2, column=ci, value=enc)
            c.fill = fill_h; c.font = font_h
            c.alignment = alin_c; c.border = borde
            ws.column_dimensions[get_column_letter(ci)].width = 20
        ws.row_dimensions[2].height = 22

        # Datos
        for ri, fila in enumerate(filas, 3):
            fill = fill_p if ri % 2 == 0 else fill_i
            vals = list(fila) if not isinstance(fila, list) else fila
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill = fill; c.font = font_d
                c.alignment = alin_c; c.border = borde

        ws.freeze_panes = "A3"

    def _estilizar_hoja_datos(self, ws, dataframe):
        """Escribe el DataFrame con formato profesional en Excel."""
        # Encabezados
        fill_header = PatternFill("solid", fgColor="0d0d1a")
        font_header = Font(bold=True, color="06d6a0", size=11)
        alin_center = Alignment(horizontal="center", vertical="center")
        borde = Border(
            bottom=Side(style="thin", color="2d2d6b"),
            right=Side(style="thin", color="2d2d6b")
        )

        for col_idx, col_name in enumerate(dataframe.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
            cell.fill      = fill_header
            cell.font      = font_header
            cell.alignment = alin_center
            cell.border    = borde
            ws.column_dimensions[get_column_letter(col_idx)].width = 18

        ws.row_dimensions[1].height = 24

        # Datos con filas alternadas
        fill_par  = PatternFill("solid", fgColor="13132b")
        fill_impar = PatternFill("solid", fgColor="1a1a3a")
        font_data  = Font(color="eaeaea", size=10)

        for row_idx, row in enumerate(dataframe.itertuples(index=False), 2):
            fill = fill_par if row_idx % 2 == 0 else fill_impar
            for col_idx, val in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill      = fill
                cell.font      = font_data
                cell.alignment = alin_center
                cell.border    = borde

        ws.freeze_panes = "A2"

    def _escribir_datos_grafico(self, ws, sub, col_x, col_y, titulo):
        """Escribe los datos del gráfico con estilo."""
        fill_h = PatternFill("solid", fgColor="2d2d6b")
        font_h = Font(bold=True, color="ffffff", size=11)
        alin   = Alignment(horizontal="center")

        ws.append([col_x.upper(), col_y.upper()])
        ws["A1"].fill = fill_h
        ws["A1"].font = font_h
        ws["A1"].alignment = alin
        ws["B1"].fill = fill_h
        ws["B1"].font = font_h
        ws["B1"].alignment = alin

        for fila in sub.itertuples(index=False):
            ws.append([str(fila[0]), fila[1]])

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 16

    def _escribir_estadisticas(self, ws):
        """Hoja de resumen estadístico de todas las columnas numéricas."""
        cols_num = [c for c in self.df.columns
                    if pd.api.types.is_numeric_dtype(self.df[c])]

        fill_h = PatternFill("solid", fgColor="0d0d1a")
        font_h = Font(bold=True, color="06d6a0", size=11)
        alin   = Alignment(horizontal="center")

        encabezados = ["COLUMNA", "REGISTROS", "MEDIA", "MEDIANA",
                       "MÍNIMO", "MÁXIMO", "DESV.STD", "SUMA"]
        for col_idx, enc in enumerate(encabezados, 1):
            cell = ws.cell(row=1, column=col_idx, value=enc)
            cell.fill = fill_h
            cell.font = font_h
            cell.alignment = alin
            ws.column_dimensions[get_column_letter(col_idx)].width = 16

        fill_alt = [PatternFill("solid", fgColor="13132b"),
                    PatternFill("solid", fgColor="1a1a3a")]
        font_d = Font(color="eaeaea", size=10)

        for i, col in enumerate(cols_num, 2):
            serie = pd.to_numeric(self.df[col], errors="coerce").dropna()
            fila = [
                col.upper(),
                len(serie),
                round(serie.mean(), 3),
                round(serie.median(), 3),
                round(serie.min(), 3),
                round(serie.max(), 3),
                round(serie.std(), 3),
                round(serie.sum(), 3),
            ]
            for col_idx, val in enumerate(fila, 1):
                cell = ws.cell(row=i, column=col_idx, value=val)
                cell.fill = fill_alt[(i) % 2]
                cell.font = font_d
                cell.alignment = alin

        if not cols_num:
            ws["A2"] = "No hay columnas numéricas en el dataset."
            ws["A2"].font = Font(color="ffbe0b", italic=True)


# ==========================================================
# UI PRINCIPAL
# ==========================================================

root = tk.Tk()
root.title("CSV Data Analyzer PRO")
root.geometry("1100x750")
root.configure(bg="#1e1e1e")

style = ttk.Style()
style.theme_use("clam")
style.configure(
    "Treeview",
    background="#2b2b2b", foreground="white",
    fieldbackground="#2b2b2b", rowheight=30,
    font=("Segoe UI", 10)
)
style.configure(
    "Treeview.Heading",
    background="#111111", foreground="white",
    font=("Segoe UI", 10, "bold")
)
style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=6)

main_frame = ttk.Frame(root, padding=15)
main_frame.pack(fill="both", expand=True)

tk.Label(
    main_frame, text="Jeanda CSV Data Analyzer PRO",
    bg="#1e1e1e", fg="white", font=("Segoe UI", 20, "bold")
).pack(pady=10)

# Botones principales
top_frame = ttk.Frame(main_frame)
top_frame.pack(fill="x", pady=10)

ttk.Button(top_frame, text="Cargar archivo",  command=cargar_csv).pack(side="left", padx=5)
ttk.Button(top_frame, text="Validar",         command=validar).pack(side="left", padx=5)
ttk.Button(top_frame, text="Estadísticas",    command=estadisticas).pack(side="left", padx=5)
ttk.Button(top_frame, text="Chart Studio",    command=mostrar_grafico).pack(side="left", padx=5)
ttk.Button(top_frame, text="Exportar",        command=exportar).pack(side="right", padx=5)

# Búsqueda
search_frame = ttk.LabelFrame(main_frame, text="Búsqueda", padding=10)
search_frame.pack(fill="x", pady=(6, 0))

ttk.Label(search_frame, text="Columna:").grid(row=0, column=0, padx=5)
combo_buscar_col = ttk.Combobox(search_frame, state="readonly", width=18)
combo_buscar_col.grid(row=0, column=1, padx=5)
combo_buscar_col.bind("<<ComboboxSelected>>", buscar_tiempo_real)

ttk.Label(search_frame, text="Buscar:").grid(row=0, column=2, padx=5)
entry_buscar = ttk.Entry(search_frame, width=30)
entry_buscar.grid(row=0, column=3, padx=5)
entry_buscar.bind("<KeyRelease>", buscar_tiempo_real)

# Filtro numérico
filter_frame = ttk.LabelFrame(main_frame, text="Filtro numérico", padding=10)
filter_frame.pack(fill="x", pady=(6, 0))

ttk.Label(filter_frame, text="Columna:").grid(row=0, column=0, padx=5)
combo_filtro_col = ttk.Combobox(filter_frame, state="readonly", width=18)
combo_filtro_col.grid(row=0, column=1, padx=5)
combo_filtro_col.bind("<<ComboboxSelected>>", filtrar_tiempo_real)

ttk.Label(filter_frame, text="Mayor o igual a:").grid(row=0, column=2, padx=5)
entry_filtro_num = ttk.Entry(filter_frame, width=16)
entry_filtro_num.grid(row=0, column=3, padx=5)
entry_filtro_num.bind("<KeyRelease>", filtrar_tiempo_real)

# Pegar datos
paste_frame = ttk.LabelFrame(main_frame, text="Pegar datos CSV", padding=10)
paste_frame.pack(fill="x", pady=(6, 0))

input_text = tk.Text(
    paste_frame, height=5,
    bg="#2b2b2b", fg="white",
    insertbackground="white",
    font=("Consolas", 10)
)
input_text.pack(fill="x", pady=5)
ttk.Button(
    paste_frame, text="Cargar datos pegados",
    command=cargar_desde_texto
).pack()

# Tabla
table_frame = ttk.Frame(main_frame)
table_frame.pack(fill="both", expand=True, pady=10)

tree = ttk.Treeview(table_frame)
tree.pack(side="left", fill="both", expand=True)

scroll_y = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
scroll_y.pack(side="right", fill="y")
tree.configure(yscrollcommand=scroll_y.set)

scroll_x = ttk.Scrollbar(main_frame, orient="horizontal", command=tree.xview)
scroll_x.pack(fill="x")
tree.configure(xscrollcommand=scroll_x.set)

root.mainloop()